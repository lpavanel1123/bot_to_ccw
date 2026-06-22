"""
Lê o XLS do ccwbot e extrai a data mais distante em 'Estimated Delivery Date'.
"""
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl


def _find_header_row(ws) -> Optional[int]:
    """Localiza a linha que contém os cabeçalhos das colunas de linha de pedido."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Estimated Delivery Date":
                return cell.row
    return None


def _find_col(ws, header_row: int, column_name: str) -> Optional[str]:
    """Retorna a letra da coluna onde está o cabeçalho informado."""
    for cell in ws[header_row]:
        if cell.value == column_name:
            return cell.column_letter
    return None


def _parse_date(value) -> Optional[datetime]:
    if not value or str(value).strip() == "":
        return None
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def latest_estimated_delivery(file_path: str) -> Optional[datetime]:
    """
    Abre o XLS/XLSX do ccwbot e retorna a data mais distante em
    'Estimated Delivery Date' entre todas as linhas de pedido.
    """
    path = Path(file_path)

    # Arquivo vem com extensão .xls mas é OOXML (ZIP) — passa como BytesIO
    # para openpyxl não rejeitar pela extensão
    wb = openpyxl.load_workbook(BytesIO(path.read_bytes()), data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("Coluna 'Estimated Delivery Date' não encontrada no arquivo.")

    col_letter = _find_col(ws, header_row, "Estimated Delivery Date")
    if col_letter is None:
        raise ValueError("Coluna 'Estimated Delivery Date' não encontrada no cabeçalho.")

    dates = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        cell = ws[f"{col_letter}{row[0].row}"]
        parsed = _parse_date(cell.value)
        if parsed:
            dates.append(parsed)

    return max(dates) if dates else None


def parse_order_lines(file_path: str) -> dict:
    """
    Parse every product line from the CCW XLS and return per-line lead time data.

    Returns:
        {
            "lines": [{"part_number", "qty", "estimated_delivery", "lead_time_days"}],
            "max_estimated_delivery": "YYYY-MM-DD" | None,
        }
    Lead time is calculated as (estimated_delivery - today).days, clamped to >= 0.
    """
    from datetime import date as _date
    today = _date.today()

    path = Path(file_path)
    wb   = openpyxl.load_workbook(BytesIO(path.read_bytes()), data_only=True)
    ws   = wb.active

    header_row = _find_header_row(ws)
    if header_row is None:
        return {"lines": [], "max_estimated_delivery": None}

    col_delivery = _find_col(ws, header_row, "Estimated Delivery Date")
    if col_delivery is None:
        return {"lines": [], "max_estimated_delivery": None}

    # Find Part Number and Qty columns by loose header match
    col_part = col_qty = None
    for cell in ws[header_row]:
        val = str(cell.value or "").strip().lower()
        if col_part is None and ("part" in val or "product" in val):
            col_part = cell.column_letter
        if col_qty is None and "qty" in val:
            col_qty = cell.column_letter

    lines    = []
    max_date = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        row_idx      = row[0].row
        delivery_val = ws[f"{col_delivery}{row_idx}"].value
        parsed       = _parse_date(delivery_val)
        if not parsed:
            continue

        part = str(ws[f"{col_part}{row_idx}"].value or "").strip() if col_part else ""
        qty  = str(ws[f"{col_qty}{row_idx}"].value  or "").strip() if col_qty  else ""

        lead_days = max((parsed.date() - today).days, 0)

        if max_date is None or parsed > max_date:
            max_date = parsed

        lines.append({
            "part_number":        part,
            "qty":                qty,
            "estimated_delivery": parsed.strftime("%Y-%m-%d"),
            "lead_time_days":     lead_days,
        })

    return {
        "lines":                  lines,
        "max_estimated_delivery": max_date.strftime("%Y-%m-%d") if max_date else None,
    }


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "order_119659099.xls"
    result = latest_estimated_delivery(path)
    print(f"Estimated Delivery Date mais distante: {result.strftime('%d-%b-%Y') if result else 'N/A'}")
